"""
Excel Export Module
===================
Generates monthly Excel reports for CA review.

OUTPUT: A styled .xlsx file with three sheets:
1. Bill Details — Every bill with vendor, amount, category, GST
2. GST Summary — Aggregated CGST/SGST/IGST totals
3. ITC Summary — Eligible vs blocked ITC with reasons

WHY openpyxl (not just pandas)?
- Pandas can write data, but openpyxl lets us STYLE it
- CAs expect formatted spreadsheets, not raw data dumps
- Headers, borders, number formatting = professional output
"""

import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings


# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
MONEY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'


def style_header_row(ws, num_cols: int):
    """Apply consistent styling to the header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)


def generate_monthly_excel(bills: list, month: int, year: int) -> str:
    """
    Generate a monthly Excel report from processed bills.

    Args:
        bills: List of Bill model objects from the database
        month: Month number (1-12)
        year: Year (e.g., 2026)

    Returns:
        Path to the generated .xlsx file
    """
    wb = Workbook()

    # ========== Sheet 1: Bill Details ==========
    ws_bills = wb.active
    ws_bills.title = "Bill Details"

    headers = [
        "S.No", "Invoice No", "Invoice Date", "Vendor Name", "Vendor GSTIN",
        "Category", "Sub-Category", "Subtotal (₹)", "CGST (₹)", "SGST (₹)",
        "IGST (₹)", "Total (₹)", "GST Rate %", "HSN Code",
        "ITC Eligible", "ITC Blocked Reason", "AI Confidence",
        "Status", "Risk Flags"
    ]
    ws_bills.append(headers)
    style_header_row(ws_bills, len(headers))

    for i, bill in enumerate(bills, 1):
        ws_bills.append([
            i,
            bill.invoice_number or "N/A",
            bill.invoice_date.strftime("%d-%m-%Y") if bill.invoice_date else "N/A",
            bill.vendor_name or "N/A",
            bill.vendor_gstin or "N/A",
            bill.final_category or bill.ai_category or "N/A",
            bill.ai_sub_category or "N/A",
            bill.subtotal,
            bill.cgst,
            bill.sgst,
            bill.igst,
            bill.total_amount,
            bill.gst_rate or 0,
            bill.hsn_code or "N/A",
            "Yes" if bill.itc_eligible else "No",
            bill.itc_blocked_reason or "-",
            f"{(bill.ai_confidence or 0):.0%}",
            bill.status.value if bill.status else "N/A",
            "Yes" if bill.needs_manual_review else "No",
        ])

    # Apply money format to amount columns
    for row in ws_bills.iter_rows(min_row=2, min_col=8, max_col=12):
        for cell in row:
            cell.number_format = MONEY_FORMAT
            cell.border = THIN_BORDER

    auto_width(ws_bills)

    # ========== Sheet 2: GST Summary ==========
    ws_gst = wb.create_sheet("GST Summary")

    gst_headers = ["Metric", "Amount (₹)"]
    ws_gst.append(gst_headers)
    style_header_row(ws_gst, len(gst_headers))

    total_cgst = sum(b.cgst for b in bills)
    total_sgst = sum(b.sgst for b in bills)
    total_igst = sum(b.igst for b in bills)
    total_gst = total_cgst + total_sgst + total_igst
    total_amount = sum(b.total_amount for b in bills)

    gst_rows = [
        ("Total Bills", len(bills)),
        ("Total Invoice Amount", total_amount),
        ("Total CGST", total_cgst),
        ("Total SGST", total_sgst),
        ("Total IGST", total_igst),
        ("Total GST (CGST + SGST + IGST)", total_gst),
    ]

    for label, value in gst_rows:
        ws_gst.append([label, value])
        ws_gst.cell(row=ws_gst.max_row, column=2).number_format = MONEY_FORMAT

    auto_width(ws_gst)

    # ========== Sheet 3: ITC Summary ==========
    ws_itc = wb.create_sheet("ITC Summary")

    itc_headers = ["Category", "ITC Eligible (₹)", "ITC Blocked (₹)", "Blocked Reason"]
    ws_itc.append(itc_headers)
    style_header_row(ws_itc, len(itc_headers))

    # Group by category
    category_itc = {}
    for bill in bills:
        cat = bill.final_category or bill.ai_category or "unclassified"
        if cat not in category_itc:
            category_itc[cat] = {"eligible": 0.0, "blocked": 0.0, "reasons": set()}

        gst_amount = (bill.cgst or 0) + (bill.sgst or 0) + (bill.igst or 0)
        if bill.itc_eligible:
            category_itc[cat]["eligible"] += gst_amount
        else:
            category_itc[cat]["blocked"] += gst_amount
            if bill.itc_blocked_reason:
                category_itc[cat]["reasons"].add(bill.itc_blocked_reason)

    for cat, data in category_itc.items():
        ws_itc.append([
            cat,
            data["eligible"],
            data["blocked"],
            "; ".join(data["reasons"]) if data["reasons"] else "-"
        ])

    # Totals row
    total_eligible = sum(d["eligible"] for d in category_itc.values())
    total_blocked = sum(d["blocked"] for d in category_itc.values())
    ws_itc.append(["TOTAL", total_eligible, total_blocked, ""])
    ws_itc.cell(row=ws_itc.max_row, column=1).font = Font(bold=True)

    auto_width(ws_itc)

    # Save file
    filename = f"GST_Report_{year}_{month:02d}.xlsx"
    filepath = os.path.join(str(settings.EXPORT_DIR), filename)
    wb.save(filepath)

    return filepath
